import openpyxl
from io import BytesIO

from django import forms
from django.contrib import admin, messages
from django.shortcuts import redirect, render
from django.urls import path
from django.utils.html import format_html
from django.utils import timezone

from .models import Tournament, TournamentFeature, TournamentRegistration
from .signals import recalculate_user_rating


class TournamentFeatureForm(forms.ModelForm):
    class Meta:
        model = TournamentFeature
        fields = '__all__'
        widgets = {
            'text': forms.TextInput(attrs={
                'maxlength': '256',
                'placeholder': 'Введите особенность (макс. 256 символов)'
            })
        }

    def clean_text(self):
        text = self.cleaned_data.get('text', '').strip()
        if len(text) > 256:
            raise forms.ValidationError("Максимальная длина 256 символов")
        if not text:
            raise forms.ValidationError("Это поле обязательно для заполнения")
        return text


class TournamentFeatureInline(admin.TabularInline):
    model = TournamentFeature
    form = TournamentFeatureForm
    extra = 1
    ordering = ('sort_order',)
    min_num = 0

    def get_queryset(self, request):
        return super().get_queryset(request).order_by('sort_order')


def _normalize_phone(phone) -> str:
    """79153530163 → +79153530163"""
    s = str(phone).strip()
    if s and not s.startswith('+'):
        s = '+' + s
    return s


def _find_col(headers_lower: list[str], candidates: list[str]) -> int | None:
    for name in candidates:
        try:
            return headers_lower.index(name)
        except ValueError:
            pass
    return None


def _to_int(val) -> int:
    """Конвертирует значение ячейки в int. Обрабатывает int, float, строки вида '25', '25.0', '25,5'."""
    if val is None:
        return 0
    if isinstance(val, bool):
        return int(val)
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip().replace(',', '.')
    if not s:
        return 0
    return int(float(s))


def _parse_excel(file) -> tuple[list[dict], list[str]]:
    """
    Парсит Excel с экспортом всех пользователей.
    data_only=True — читает кешированные результаты формул, а не сами формулы.
    Возвращает (rows, errors) — только строки с ИГР Рейт > 0 или ИГР Кил > 0.
    """
    wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
    ws = wb.active
    if ws is None:
        return [], ["Не удалось открыть лист Excel"]

    # Ищем строку с заголовками в первых 5 строках
    header_row_num = None
    hl = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        row_lower = [str(v).strip().lower() if v is not None else '' for v in row]
        if 'игр рейт' in row_lower or 'игр_рейт' in row_lower or 'points' in row_lower:
            header_row_num = i
            hl = row_lower
            break

    if header_row_num is None:
        return [], ["Не найдена строка с заголовками (ищу 'ИГР Рейт' в первых 5 строках)"]

    id_idx        = _find_col(hl, ['id'])
    phone_idx     = _find_col(hl, ['phone', 'телефон'])
    nickname_idx  = _find_col(hl, ['nickname', 'никнейм'])
    points_idx    = _find_col(hl, ['игр рейт', 'игр_рейт', 'points'])
    knockouts_idx = _find_col(hl, ['игр кил', 'игр_кил', 'knockouts'])

    errors = []
    if points_idx is None:
        errors.append("Не найдена колонка 'ИГР Рейт' (или 'points')")
    if knockouts_idx is None:
        errors.append("Не найдена колонка 'ИГР Кил' (или 'knockouts')")
    if errors:
        return [], errors

    rows = []
    for row_num, row in enumerate(ws.iter_rows(min_row=header_row_num + 1, values_only=True), start=header_row_num + 1):
        if not any(v is not None for v in row):
            continue

        try:
            points    = _to_int(row[points_idx])
            knockouts = _to_int(row[knockouts_idx])
        except (ValueError, TypeError) as e:
            errors.append(f'Строка {row_num}: не удалось прочитать ИГР Рейт/ИГР Кил — {e}')
            continue

        if points == 0 and knockouts == 0:
            continue  # не участвовал в этом турнире

        user_id  = _to_int(row[id_idx]) if id_idx is not None and row[id_idx] is not None else None
        phone    = _normalize_phone(row[phone_idx]) if phone_idx is not None and row[phone_idx] else ''
        nickname = str(row[nickname_idx]).strip().lower() if nickname_idx is not None and row[nickname_idx] else ''

        rows.append({
            'user_id': user_id or None,
            'phone': phone,
            'nickname': nickname,
            'points': points,
            'knockouts': knockouts,
        })

    return rows, errors


@admin.register(Tournament)
class TournamentAdmin(admin.ModelAdmin):
    inlines = [TournamentFeatureInline]

    list_display = ['id', 'title', 'location', 'started_at_compact', 'status_badge', 'participants_count', 'tournament_action_button']
    list_display_links = ['id', 'title']
    list_filter = ['status']
    search_fields = ['title', 'location']
    ordering = ['-started_at']

    class Media:
        css = {
            'all': ('admin/tournament.css',)
        }

    def get_fieldsets(self, request, obj=None):
        if obj:
            return (
                ('Основная информация', {
                    'fields': ('title', 'location', 'started_at', 'general_rules', 'max_participants', 'max_waitlist', 'icon')
                }),
                ('Статус', {
                    'fields': ('status_display',)
                }),
                ('Даты', {
                    'fields': ('created_at', 'updated_at')
                }),
                ('Участники', {
                    'fields': ('participants_count_display',)
                }),
            )
        return (
            ('Основная информация', {
                'fields': ('title', 'location', 'started_at', 'general_rules', 'max_participants', 'max_waitlist', 'icon')
            }),
        )

    def participants_count(self, obj):
        return obj.get_participants_count()
    participants_count.short_description = 'Участники'

    def participants_count_display(self, obj):
        count = obj.get_participants_count()
        return format_html('<strong>{}</strong> участника(-ов)', count)
    participants_count_display.short_description = 'Зарегистрировано участников'

    def get_readonly_fields(self, request, obj=None):
        if not obj:
            return []

        readonly_fields = ['created_at', 'updated_at', 'status_display', 'participants_count_display']

        if obj.status == 'INACTIVE':
            all_fields = [f.name for f in self.model._meta.fields]
            readonly_fields = all_fields + ['status_display', 'participants_count_display']

        return readonly_fields

    def get_exclude(self, request, obj=None):
        if not obj:
            return ['status']
        return []

    def started_at_compact(self, obj):
        return timezone.localtime(obj.started_at).strftime('%d.%m.%Y %H:%M')
    started_at_compact.short_description = 'Старт'
    started_at_compact.admin_order_field = 'started_at'

    def status_display(self, obj):
        return self.status_badge(obj)
    status_display.short_description = 'Текущий статус'

    def tournament_action_button(self, obj):
        if obj.status == 'IN_QUEUE':
            return format_html(
                '<a class="button" href="start_tournament/{}/" style="background: #28a745; color: white; padding: 5px 10px; text-decoration: none; border-radius: 4px; font-size: 12px;">Запустить турнир</a>',
                obj.id
            )
        elif obj.status == 'ACTIVE':
            return format_html(
                '<a class="button" href="finish_tournament/{}/" style="background: #dc3545; color: white; padding: 5px 10px; text-decoration: none; border-radius: 4px; font-size: 12px;">Завершить турнир</a>',
                obj.id
            )
        elif obj.status == 'INACTIVE':
            return format_html(
                '<a class="button" href="upload_results/{}/" style="background: #007bff; color: white; padding: 5px 10px; text-decoration: none; border-radius: 4px; font-size: 12px;">Загрузить результаты</a>',
                obj.id
            )
        return ""
    tournament_action_button.short_description = 'Действие'

    def status_badge(self, obj):
        status_config = {
            'IN_QUEUE': ('🟡', 'yellow', 'В очереди'),
            'ACTIVE': ('🟢', 'green', 'Активный'),
            'INACTIVE': ('⚫', 'gray', 'Завершен')
        }
        emoji, color, text = status_config.get(obj.status, ('⚫', 'gray', obj.status))
        return format_html(
            '{} <span style="background: {}; color: white; padding: 4px 8px; border-radius: 12px; font-size: 11px; font-weight: 500;">{}</span>',
            emoji, color, text
        )
    status_badge.short_description = 'Статус'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'start_tournament/<int:tournament_id>/',
                self.admin_site.admin_view(self.start_tournament),
                name='start_tournament'
            ),
            path(
                'finish_tournament/<int:tournament_id>/',
                self.admin_site.admin_view(self.finish_tournament),
                name='finish_tournament'
            ),
            path(
                'upload_results/<int:tournament_id>/',
                self.admin_site.admin_view(self.upload_results),
                name='upload_results'
            ),
        ]
        return custom_urls + urls

    def start_tournament(self, request, tournament_id):
        try:
            tournament = Tournament.objects.get(id=tournament_id)
            tournament.started_at = timezone.now()
            tournament.status = 'ACTIVE'
            tournament.save()
            self.message_user(request, f'Турнир "{tournament.title}" успешно запущен', messages.SUCCESS)
        except Tournament.DoesNotExist:
            self.message_user(request, 'Турнир не найден', messages.ERROR)

        return redirect('admin:tournaments_tournament_changelist')

    def finish_tournament(self, request, tournament_id):
        try:
            tournament = Tournament.objects.get(id=tournament_id)
            tournament.status = 'INACTIVE'
            tournament.save()
            self.message_user(request, f'Турнир "{tournament.title}" успешно завершен', messages.SUCCESS)
        except Tournament.DoesNotExist:
            self.message_user(request, 'Турнир не найден', messages.ERROR)

        return redirect('admin:tournaments_tournament_changelist')

    def upload_results(self, request, tournament_id):
        try:
            tournament = Tournament.objects.get(id=tournament_id)
        except Tournament.DoesNotExist:
            self.message_user(request, 'Турнир не найден', messages.ERROR)
            return redirect('admin:tournaments_tournament_changelist')

        if tournament.status != Tournament.StatusType.INACTIVE:
            self.message_user(request, 'Загрузка результатов доступна только для завершённых турниров.', messages.ERROR)
            return redirect('admin:tournaments_tournament_changelist')

        context = {
            **self.admin_site.each_context(request),
            'tournament': tournament,
            'title': f'Загрузить результаты: {tournament.title}',
            'result': None,
            'errors': [],
        }

        if request.method != 'POST' or not request.FILES.get('excel_file'):
            return render(request, 'admin/tournaments/upload_results.html', context)

        try:
            rows, parse_errors = _parse_excel(request.FILES['excel_file'])
        except Exception as e:
            self.message_user(request, f'Ошибка чтения файла: {e}', messages.ERROR)
            return render(request, 'admin/tournaments/upload_results.html', context)

        if parse_errors and not rows:
            context['errors'] = parse_errors
            return render(request, 'admin/tournaments/upload_results.html', context)

        if not rows:
            self.message_user(request, 'Нет участников с ненулевыми результатами', messages.WARNING)
            return render(request, 'admin/tournaments/upload_results.html', context)

        # Индекс регистраций по user_id, phone, nickname
        regs_by_user_id  = {}
        regs_by_phone    = {}
        regs_by_nickname = {}
        for reg in TournamentRegistration.objects.filter(tournament=tournament).select_related('user'):
            regs_by_user_id[reg.user_id] = reg
            if reg.user.phone:
                regs_by_phone[reg.user.phone.strip()] = reg
            regs_by_nickname[reg.user.nickname.strip().lower()] = reg

        from django.contrib.auth import get_user_model
        User = get_user_model()

        to_update = []
        to_create = []
        affected_user_ids = set()
        errors = list(parse_errors)

        for row in rows:
            reg = (
                regs_by_user_id.get(row['user_id'])
                or regs_by_phone.get(row['phone'])
                or regs_by_nickname.get(row['nickname'])
            )

            if reg is not None:
                reg.knockouts = row['knockouts']
                reg.points = row['points']
                reg.attended = True
                to_update.append(reg)
                affected_user_ids.add(reg.user_id)
            else:
                # Регистрации нет — ищем пользователя для авто-регистрации
                user = None
                if row['user_id']:
                    user = User.objects.filter(pk=row['user_id']).first()
                if not user and row['phone']:
                    user = User.objects.filter(phone=row['phone']).first()
                if not user and row['nickname']:
                    user = User.objects.filter(nickname__iexact=row['nickname']).first()

                if user:
                    to_create.append(TournamentRegistration(
                        tournament=tournament,
                        user=user,
                        status=TournamentRegistration.StatusType.REGISTERED,
                        points=row['points'],
                        knockouts=row['knockouts'],
                        attended=True,
                    ))
                    affected_user_ids.add(user.id)
                else:
                    identifier = row['phone'] or row['nickname'] or f"ID={row['user_id']}"
                    errors.append(f'Пользователь не найден: «{identifier}» (ИГР Рейт={row["points"]}, ИГР Кил={row["knockouts"]})')

        # Сброс всех регистраций турнира перед применением нового файла,
        # чтобы пользователи из предыдущей загрузки не сохраняли старые очки.
        all_reg_user_ids = set(
            TournamentRegistration.objects.filter(tournament=tournament).values_list('user_id', flat=True)
        )
        TournamentRegistration.objects.filter(tournament=tournament).update(
            points=0, knockouts=0, attended=False
        )

        if to_update:
            TournamentRegistration.objects.bulk_update(to_update, ['knockouts', 'points', 'attended'])

        if to_create:
            TournamentRegistration.objects.bulk_create(to_create)

        # Пересчитываем всех, у кого были или появились очки
        for user_id in all_reg_user_ids | affected_user_ids:
            recalculate_user_rating(user_id)

        context['result'] = {'updated': len(to_update), 'created': len(to_create)}
        context['errors'] = errors

        if to_update or to_create:
            self.message_user(
                request,
                f'Обновлено {len(to_update)} регистраций, создано {len(to_create)} новых.',
                messages.SUCCESS,
            )

        return render(request, 'admin/tournaments/upload_results.html', context)

    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        if object_id:
            try:
                obj = Tournament.objects.get(id=object_id)
                if obj.status == 'INACTIVE':
                    extra_context = extra_context or {}
                    extra_context['readonly_message'] = 'Этот турнир завершен и не может быть изменен.'
            except Tournament.DoesNotExist:
                pass

        return super().changeform_view(request, object_id, form_url, extra_context)


@admin.register(TournamentRegistration)
class TournamentRegistrationAdmin(admin.ModelAdmin):
    list_display = ['id', 'tournament', 'tournament_status', 'user_nickname', 'user_username', 'user_phone', 'knockouts', 'points', 'status_badge', 'created_at_compact']
    list_filter = ['status', 'tournament', 'created_at']
    search_fields = ['user__username', 'user__nickname', 'tournament__title']
    readonly_fields = ['created_at', 'updated_at']
    exclude = ['attended']
    list_select_related = ['user', 'tournament']

    @admin.display(description='Статус')
    def status_badge(self, obj):
        status_config = {
            'REGISTERED': ('🟢', 'green', 'Зарегистрирован'),
            'WAITLIST': ('🟡', 'yellow', 'В листе ожидания'),
            'CANCELLED': ('🔴', 'red', 'Отменена'),
        }
        emoji, color, text = status_config.get(obj.status, ('⚫', 'gray', obj.status))
        return format_html(
            '<span style="white-space: nowrap;">{} '
            '<span style="background: {}; color: white; padding: 4px 8px; border-radius: 12px; font-size: 11px; font-weight: 500;">{}</span>'
            '</span>',
            emoji, color, text
        )

    @admin.display(ordering="tournament__status", description="Tournament status")
    def tournament_status(self, obj):
        return obj.tournament.status

    @admin.display(ordering="user__nickname", description="Nickname")
    def user_nickname(self, obj):
        return obj.user.nickname

    @admin.display(ordering="user__username", description="Username")
    def user_username(self, obj):
        return obj.user.username

    @admin.display(ordering="user__phone", description="Phone")
    def user_phone(self, obj):
        return obj.user.phone

    @admin.display(ordering='created_at', description='Зарегистрирован')
    def created_at_compact(self, obj):
        return timezone.localtime(obj.created_at).strftime('%d.%m.%Y %H:%M')
